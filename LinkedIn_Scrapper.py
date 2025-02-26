from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()
ws = wb.active
ws.append(["Name", "Next Line After Tag"])

neutral_fill = PatternFill(start_color='E0E0E0', fill_type='solid')
for cell in ws[1]:  
    cell.fill = neutral_fill

os.system("cls")

EMAIL = input ("\nplease Enter LinkedIn Username: ")
PASSWORD = input ("\nplease Enter LinkedIn Password: ")

company_url = input("\n\nPlease Enter Main Page of Company\nFor Example ---> https://www.linkedin.com/company/google/people\n\n---> ")

driver = webdriver.Chrome()

try:
    driver.get('https://www.linkedin.com/login')
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'username')))

    driver.find_element(By.ID, 'username').send_keys(EMAIL)
    driver.find_element(By.ID, 'password').send_keys(PASSWORD)
    driver.find_element(By.XPATH, '//button[@type="submit"]').click()

    time.sleep(10) 

    driver.get(company_url)
    time.sleep(5)

    last_height = driver.execute_script("return document.body.scrollHeight")
    
    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(3)
        new_height = driver.execute_script("return document.body.scrollHeight")

        if new_height == last_height:
            break
        last_height = new_height

    with open('CleanData.txt', 'w', encoding='utf-8') as file:
        file.write(driver.page_source)

    print("HTML source has been saved to 'CleanData.txt'")

except Exception as e:
    print(f"An error occurred: {e}")

finally:
    driver.quit()

with open('CleanData.txt', 'r', encoding='utf-8') as file:
    lines = file.readlines()

    for i in range(len(lines)):
        line = lines[i]

        match_profile = re.search(r'View (.*?)’s profile', line)
        if match_profile:
            name = match_profile.group(1).strip()
            ws.append([name, ""])  

        if '<!----><!----></div>' in line and i + 4 < len(lines):
            next_line = lines[i + 4].strip()
            ws.append(["", next_line])  

wb.save("output.xlsx")
print("Data extraction and Excel file creation completed successfully.")
